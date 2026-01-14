"""
Export helper for handling credential exports in various formats
"""
import io
import csv
import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import Response
from src.utils.senitization import sanitize_for_excel


class ExportHelper:
    """Helper class for exporting credentials in various formats"""
    
    HEADERS = ["Domain", "URL", "Username", "Password", "Is Admin", "Is Accessed", "Is Checked", "Status"]
    HEADER_COLOR = "667eea"
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename_prefix: str = "credentials") -> Response:
        """Export data to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Credentials"
        
        # Add headers with styling
        header_fill = PatternFill(start_color=ExportHelper.HEADER_COLOR, end_color=ExportHelper.HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(ExportHelper.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        row_num = 2
        for item in data:
            ws.cell(row=row_num, column=1, value=sanitize_for_excel(item.get("domain", "")))
            ws.cell(row=row_num, column=2, value=sanitize_for_excel(item.get("url", "")))
            ws.cell(row=row_num, column=3, value=sanitize_for_excel(item.get("username", item.get("user", ""))))
            ws.cell(row=row_num, column=4, value=sanitize_for_excel(item.get("password", "")))
            ws.cell(row=row_num, column=5, value=item.get("is_admin", "No"))
            ws.cell(row=row_num, column=6, value=item.get("is_accessed", "No"))
            ws.cell(row=row_num, column=7, value=item.get("is_checked", "No"))
            ws.cell(row=row_num, column=8, value=sanitize_for_excel(item.get("status", "N/A")))
            row_num += 1
        
        # Adjust column widths
        ExportHelper._adjust_column_widths(ws)
        
        # Prepare response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename_prefix: str = "credentials") -> Response:
        """Export data to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(ExportHelper.HEADERS)
        
        # Write data
        for item in data:
            writer.writerow([
                item.get("domain", ""),
                item.get("url", ""),
                item.get("username", item.get("user", "")),
                item.get("password", ""),
                item.get("is_admin", "No"),
                item.get("is_accessed", "No"),
                item.get("is_checked", "No"),
                item.get("status", "N/A")
            ])
        
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.csv"
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @staticmethod
    def export_to_txt(data: List[Dict[str, Any]], filename_prefix: str = "credentials") -> Response:
        """Export data to TXT format"""
        output = io.StringIO()
        
        for item in data:
            output.write(f"Domain: {item.get('domain', '')}\n")
            output.write(f"URL: {item.get('url', '')}\n")
            output.write(f"Username: {item.get('username', item.get('user', ''))}\n")
            output.write(f"Password: {item.get('password', '')}\n")
            output.write(f"Is Admin: {item.get('is_admin', 'No')}\n")
            output.write(f"Is Accessed: {item.get('is_accessed', 'No')}\n")
            output.write(f"Is Checked: {item.get('is_checked', 'No')}\n")
            output.write(f"Status: {item.get('status', 'N/A')}\n")
            output.write("-" * 50 + "\n\n")
        
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.txt"
        
        return Response(
            content=output.getvalue(),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], filename_prefix: str = "credentials") -> Response:
        """Export data to JSON format"""
        output = json.dumps(data, indent=2, ensure_ascii=False)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.json"
        
        return Response(
            content=output,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @staticmethod
    def _adjust_column_widths(worksheet):
        """Adjust column widths for Excel worksheet"""
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width
    
    @staticmethod
    def prepare_export_data(credentials, domains=None) -> List[Dict[str, Any]]:
        """Prepare data for export from credentials and domains"""
        if domains:
            # Create domain map for quick lookup
            domain_map = {domain.id: domain for domain in domains}
        
        export_data = []
        for cred in credentials:
            if domains:
                domain = domain_map.get(cred.domain_id)
                domain_name = domain.domain if domain else "Unknown"
            else:
                domain_name = getattr(cred, 'domain_name', 'Unknown')
            
            export_data.append({
                "domain": domain_name,
                "url": cred.url,
                "username": cred.user,
                "password": cred.password,
                "is_admin": "Yes" if cred.is_admin else "No",
                "is_accessed": "Yes" if cred.is_accessed else "No",
                "is_checked": "Yes" if cred.is_checked else "No",
                "status": cred.status.name if cred.status else "N/A"
            })
        
        return export_data